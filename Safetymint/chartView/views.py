from calendar import calendar, month, month_abbr
import re
from django.http import HttpResponseRedirect
from django.shortcuts import render
from .forms import UploadFileForm
import openpyxl

from datetime import datetime



def upload_file_view(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        # if form.is_valid():
        uploaded_file=request.FILES['xlfile']
        print(uploaded_file.name)
        data_set=handle_uploaded_file(uploaded_file)
        dates={}
        temp=list(data_set.keys())
        temp.sort(key=lambda date:datetime.strptime(date,"%b-%y"))
        for i in temp:
            dates[i]=data_set[i]
    
        # dates.sort(key=lambda date:datetime.strptime(date,"%b-%y"))
        print(dates)
        
        # data_count=[]
        # for i in range(len(dates)):
        #     data_count.append(data_set[dates[i]])
        # print(data_count)
        # x=zip(dates,data_count)
        # print(x)
        
        
        
        return render(request,"charts.html",{'x':dates})
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})



# def Month_Name_Using_Date(date):
#     d = datetime.strptime(date,'%d-%b-%y')
#     Month_name=d.strftime('%b-%y')
#     return Month_name

def handle_uploaded_file(file):
    
    wb=openpyxl.load_workbook(file)
    sheet_obj=wb.active
    
    data_dict={}
    rowCount=sheet_obj.max_row
    colCount=sheet_obj.max_column
    t=["Minor","Serious","Critical"]

    for i in range(2,rowCount+1):
        cell_date=datetime.strftime(sheet_obj.cell(row=i,column=2).value,"%b-%y")
        cell_type=sheet_obj.cell(row=i,column=3).value
        if cell_date not in data_dict:
            data_dict[cell_date]=[0,0,0]
        
        data_dict[cell_date][t.index(cell_type)]+=1

        

    return data_dict

# def chartdisp(li):
    
#     return